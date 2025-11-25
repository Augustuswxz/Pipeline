import os
import openpyxl
from langchain.tools import tool

@tool
def data_alignment_tool(filename: str, threshold: float) -> str:
    """
    对齐指定Excel文件中的数据。
    
    读取文件中的Sheet1和Sheet2，基于'绝对距离'计算相对距离，
    使用指定的【误差阈值】进行对齐匹配，并将结果写入Sheet3。
    
    Args:
        filename: Excel文件的路径（例如 "test1.xlsx" 或绝对路径）
        threshold: 用于判断数据对齐的绝对误差阈值（例如 0.2, 0.5）
        
    Returns:
        str: 执行结果的消息
    """
    
    # 检查文件是否存在
    base_dir = os.getcwd()
    file_path = os.path.join(base_dir, "UploadedFiles", os.path.basename(filename))
    if not os.path.exists(file_path):
        return f"错误：找不到文件 {filename}"

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path)

        # ================= 处理 Sheet1 =================
        sheet1 = workbook['Sheet1']
        titles1 = list(sheet1[1])
        dis_loc1 = 64
        values1 = []
        for title in titles1:
            values1.append(title.value)
        
        # 寻找“绝对距离”列
        found_col1 = False
        for value in values1:
            dis_loc1 = dis_loc1 + 1
            if value == '绝对距离':
                found_col1 = True
                break
        
        if not found_col1:
            return "错误：在 Sheet1 中未找到 '绝对距离' 列"

        dis_loc1 = chr(dis_loc1)
        dis_area1 = list(sheet1[dis_loc1])
        dis_area1.pop(0) # 去掉表头
        dis1 = []
        for cell in dis_area1:
            if cell.value is not None:
                dis1.append(float(cell.value))
            else:
                dis1.append(0.0)

        Redis1 = []
        for i in range(len(dis1)):
            if i + 1 < len(dis1):
                Redis1.append(round(dis1[i + 1] - dis1[i], 4))

        # ================= 处理 Sheet2 =================
        sheet2 = workbook['Sheet2']
        titles2 = list(sheet2[1])
        dis_loc2 = 64
        values2 = []
        for title in titles2:
            values2.append(title.value)
            
        found_col2 = False
        for value in values2:
            dis_loc2 = dis_loc2 + 1
            if value == '绝对距离':
                found_col2 = True
                break
        
        if not found_col2:
            return "错误：在 Sheet2 中未找到 '绝对距离' 列"

        dis_loc2 = chr(dis_loc2)
        dis_area2 = list(sheet2[dis_loc2])
        dis_area2.pop(0)
        dis2 = []
        for cell in dis_area2:
            if cell.value is not None:
                dis2.append(float(cell.value))
            else:
                dis2.append(0.0)

        Redis2 = []
        for i in range(len(dis2)):
            if i + 1 < len(dis2):
                Redis2.append(round(dis2[i + 1] - dis2[i], 4))

        # ================= 准备 Sheet3 输出 =================
        # 如果 Sheet3 不存在则创建，存在则获取（这里假设是新建或追加）
        if 'Sheet3' in workbook.sheetnames:
            sheet3 = workbook['Sheet3']
        else:
            sheet3 = workbook.create_sheet('Sheet3')
            
        # 每次调用建议清空旧数据或仅追加，这里保持原逻辑直接 append
        sheet3.append(['上游环焊缝编号', '绝对距离', '相对距离', '上游环焊缝编号', '绝对距离', '相对距离'])

        # ================= 核心对齐逻辑 =================
        weld1 = 1
        weld2 = 1
        begin1 = 0
        begin2 = 0
        num1 = 0
        num2 = 0
        count = 0
        bad_count = 0
        
        # 第一阶段：寻找起始匹配点
        while num1 < len(Redis1) and num2 < len(Redis2):
            # 【修改点】这里使用了传入的 threshold 参数替换了硬编码的 0.2
            if abs(Redis1[num1] - Redis2[num2]) < threshold:
                if count == 0:
                    begin1 = num1
                    begin2 = num2
                count = count + 1
                num1 = num1 + 1
                num2 = num2 + 1
            else:  # 不匹配的话重新开始计数
                num1 = num1 + 1
                num2 = 0
                count = 0
                bad_count = bad_count + 1
            
            if count > 3:  # 四个连续匹配，认为找到了起始点
                if begin1 == 0:
                    sheet3.append(['', '', '', str(weld2), str(dis2[weld2 - 1]), '0'])
                    weld2 = weld2 + 1
                    while weld2 - 2 < begin2:
                        if weld2 - 2 == begin2 - 1:
                            sheet3.append([
                                str(weld1), str(dis1[weld1 - 1]), '0',
                                str(weld2), str(dis2[weld2 - 1]), str(Redis2[weld2 - 2])
                            ])
                            weld1 = weld1 + 1
                        else:
                            sheet3.append(['', '', '', str(weld2), str(dis2[weld2 - 1]), str(Redis2[weld2 - 2])])
                        weld2 = weld2 + 1
                    
                    while weld2 - 2 < num2:
                        sheet3.append([
                            str(weld1), str(dis1[weld1 - 1]), str(Redis1[weld2 - 2 - (begin2 - begin1)]),
                            str(weld2), str(dis2[weld2 - 1]), str(Redis2[weld2 - 2])
                        ])
                        weld1 = weld1 + 1
                        weld2 = weld2 + 1
                else:
                    sheet3.append([str(weld1), str(dis1[weld1 - 1]), '0', '', '', ''])
                    weld1 = weld1 + 1
                    while weld1 - 2 < begin1:
                        if weld1 - 2 == begin1 - 1:
                            sheet3.append([
                                str(weld1), str(dis1[weld1 - 1]), str(Redis1[weld1 - 2]),
                                str(weld2), str(dis2[weld2 - 1]), '0'
                            ])
                            weld2 = weld2 + 1
                        else:
                            sheet3.append([str(weld1), str(dis1[weld1 - 1]), str(Redis1[weld1 - 2]), '', '', ''])
                        weld1 = weld1 + 1
                    
                    while weld1 - 2 < num1:
                        sheet3.append([
                            str(weld1), str(dis1[weld1 - 1]), str(Redis1[weld1 - 2]),
                            str(weld2), str(dis2[weld2 - 1]), str(Redis2[weld1 - 2 - (begin1 - begin2)])
                        ])
                        weld1 = weld1 + 1
                        weld2 = weld2 + 1
                break
            
            if bad_count > 10:
                num1 = 0
                num2 = num2 + 1
                bad_count = 0
                continue

        # 第二阶段：后续对齐
        # 注意：这里如果你想把 threshold 也应用到下面的逻辑，可以替换 0.06
        # 但根据通常逻辑，0.2 是绝对误差(值)，0.06 是相对误差(百分比)
        # 这里暂且保留原逻辑，只替换起始点的绝对误差
        if num1 < len(Redis1) and num2 < len(Redis2):
            now_align1 = Redis1[num1]
            now_align2 = Redis2[num2]
            
            while num1 < len(Redis1) and num2 < len(Redis2):
                # 相对误差判断
                if abs(now_align1 - now_align2)/now_align1 < 0.06 or abs(now_align1 - now_align2)/now_align2 < 0.06:
                    sheet3.append([
                        str(weld1), str(dis1[num1 + 1]), str(Redis1[num1]),
                        str(weld2), str(dis2[num2 + 1]), str(Redis2[num2])
                    ])
                    num1 = num1 + 1
                    num2 = num2 + 1
                    weld1 = weld1 + 1
                    weld2 = weld2 + 1
                    if num1 < len(Redis1) and num2 < len(Redis2):
                        now_align1 = Redis1[num1]
                        now_align2 = Redis2[num2]
                elif now_align1 < now_align2:
                    sheet3.append([str(weld1), str(dis1[num1 + 1]), str(Redis1[num1]), '', '', ''])
                    num1 = num1 + 1
                    weld1 = weld1 + 1
                    if num1 < len(Redis1):
                        now_align1 = now_align1 + Redis1[num1]
                else:
                    sheet3.append(['', '', '', str(weld2), str(dis2[num2 + 1]), str(Redis2[num2])])
                    num2 = num2 + 1
                    weld2 = weld2 + 1
                    if num2 < len(Redis2):
                        now_align2 = now_align2 + Redis2[num2]

        # 保存文件
        save_file_path = os.path.join(base_dir, "AlignedFiles", os.path.basename(filename))
        workbook.save(save_file_path)
        return f"成功：已完成文件 {filename} 的数据对齐，阈值设定为 {threshold}，结果已保存至 Sheet3。"

    except Exception as e:
        return f"执行对齐工具时发生异常: {str(e)}"